# scheduler.py
#
# Build your bootcamp scheduler here.
#
# Work through the challenges in order:
#   Run /next in Claude Code to see what to do next.
#   Run /spec to write a specification for the next function.
#   Run /build to implement from your spec.
#   Run python3 -m pytest tests/ -v to verify what you've built.
#
# Each challenge adds new functions — don't remove old ones.

from datetime import date, timedelta
from openpyxl import load_workbook
from ortools.sat.python import cp_model
import holidays as _holidays_lib

_DAY_ABBREV = {
    "mon": 0, "tue": 1, "wed": 2, "thu": 3, "fri": 4, "sat": 5, "sun": 6,
}

_METADATA_FIELDS = {
    "home location": "home_location",
    "french": "french_speaking",
    "max/week": "max_per_week",
}


def parse_availability(filepath: str):
    """Parse a trainer-availability workbook.

    Returns either a 2-tuple ``(dates, trainers)`` for the basic format, or a
    4-tuple ``(dates, trainers, trainer_info, weekly_caps)`` when the workbook
    has extended columns (``Home Location``, ``French``, ``Max/Week``).

    Detection is by header content: any of the three metadata column names
    triggers the extended return.

    Raises:
        ValueError: If a date header is malformed or duplicated.
    """
    workbook = load_workbook(filepath, read_only=True, data_only=True)
    worksheet = workbook["Availability"] if "Availability" in workbook.sheetnames else workbook.active

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return ([], {})

    header = rows[0]
    date_cols, metadata_cols = _split_header(header)
    sorted_dates = sorted(d for _, d in date_cols)

    # Caps row: present iff the row after the header has an empty Name cell.
    caps_row = None
    trainer_rows_start = 1
    if len(rows) > 1:
        candidate = rows[1]
        name_cell = candidate[0] if len(candidate) > 0 else None
        if name_cell is None or not str(name_cell).strip():
            caps_row = candidate
            trainer_rows_start = 2

    weekly_caps = _parse_caps_row(caps_row, date_cols) if caps_row else {}

    trainers: dict[str, dict[date, bool]] = {}
    trainer_info: dict[str, dict] = {}
    has_metadata = bool(metadata_cols)

    for row in rows[trainer_rows_start:]:
        if not row:
            continue
        raw_name = row[0] if len(row) > 0 else None
        if raw_name is None:
            continue
        name = str(raw_name).strip()
        if not name:
            continue

        availability = {
            d: _is_available(row[col_idx] if col_idx < len(row) else None)
            for col_idx, d in date_cols
        }
        trainers[name] = {d: availability[d] for d in sorted_dates}

        if has_metadata:
            trainer_info[name] = _parse_trainer_metadata(row, metadata_cols)

    if has_metadata:
        return (sorted_dates, trainers, trainer_info, weekly_caps)
    return (sorted_dates, trainers)


def _split_header(header: tuple) -> tuple[list[tuple[int, date]], dict[int, str]]:
    """Walk the header row and classify each column.

    Returns ``(date_cols, metadata_cols)`` where ``date_cols`` is a list of
    ``(column_index, date)`` for date columns and ``metadata_cols`` is a
    ``{column_index: normalized_field_name}`` for metadata columns.

    Raises:
        ValueError: On malformed or duplicate date headers.
    """
    date_cols: list[tuple[int, date]] = []
    metadata_cols: dict[int, str] = {}
    seen_dates: set[date] = set()

    for col_idx, cell in enumerate(header):
        if col_idx == 0 or cell is None:
            continue
        text = str(cell).strip()
        if not text:
            continue

        # Try date first
        try:
            parsed_date = date.fromisoformat(text)
        except ValueError:
            field = _METADATA_FIELDS.get(text.lower())
            if field is not None:
                metadata_cols[col_idx] = field
            continue

        if parsed_date in seen_dates:
            raise ValueError(f"Duplicate date in header: {parsed_date.isoformat()}")
        seen_dates.add(parsed_date)
        date_cols.append((col_idx, parsed_date))

    return date_cols, metadata_cols


def _parse_caps_row(
    caps_row: tuple,
    date_cols: list[tuple[int, date]],
) -> dict[tuple[int, int], int]:
    """Build ``{(iso_year, iso_week): cap}`` from a caps row, inspecting only Monday columns."""
    result: dict[tuple[int, int], int] = {}
    for col_idx, d in date_cols:
        if d.weekday() != 0:  # Monday only
            continue
        if col_idx >= len(caps_row):
            continue
        cell = caps_row[col_idx]
        if cell is None:
            continue
        try:
            value = int(cell)
        except (TypeError, ValueError):
            continue
        iso_year, iso_week, _ = d.isocalendar()
        result[(iso_year, iso_week)] = value
    return result


_CITY_TO_ISO: dict[str, str] = {
    # United Kingdom
    "London": "GB", "Manchester": "GB", "Bristol": "GB",
    "Birmingham": "GB", "Glasgow": "GB", "Edinburgh": "GB", "Liverpool": "GB",
    # France
    "Paris": "FR", "Lyon": "FR", "Marseille": "FR", "Toulouse": "FR", "Nice": "FR",
    # Netherlands / Belgium / Germany / Switzerland / Austria
    "Amsterdam": "NL", "Rotterdam": "NL", "The Hague": "NL", "Utrecht": "NL",
    "Brussels": "BE", "Antwerp": "BE",
    "Berlin": "DE", "Munich": "DE", "Frankfurt": "DE", "Hamburg": "DE",
    "Zurich": "CH", "Geneva": "CH", "Basel": "CH",
    "Vienna": "AT",
    # Iberia / Italy / Nordics / Other Europe
    "Madrid": "ES", "Barcelona": "ES", "Seville": "ES",
    "Lisbon": "PT", "Porto": "PT",
    "Rome": "IT", "Milan": "IT", "Turin": "IT",
    "Stockholm": "SE", "Gothenburg": "SE",
    "Copenhagen": "DK", "Oslo": "NO", "Helsinki": "FI",
    "Dublin": "IE", "Warsaw": "PL", "Prague": "CZ", "Budapest": "HU",
    # Long-haul (so apply_bank_holidays still works for them)
    "New York": "US", "Los Angeles": "US", "San Francisco": "US", "Chicago": "US",
    "Mumbai": "IN", "Delhi": "IN", "Bangalore": "IN",
    "Singapore": "SG",
    "Sydney": "AU", "Melbourne": "AU",
    "Tokyo": "JP",
    "Hong Kong": "HK",
}


def apply_bank_holidays(
    trainers: dict[str, dict[date, bool]],
    trainer_info: dict[str, dict] | None,
) -> None:
    """Mark trainer availability ``False`` on their home country's public holidays.

    Modifies ``trainers`` in place. No-op when ``trainer_info`` is missing,
    empty, or a trainer's ``home_location`` is not in :data:`_CITY_TO_ISO`.
    """
    if not trainer_info:
        return

    holiday_cache: dict[tuple[str, frozenset[int]], object] = {}

    for name, info in trainer_info.items():
        if name not in trainers:
            continue
        city = (info or {}).get("home_location")
        if not city:
            continue
        iso = _CITY_TO_ISO.get(str(city).strip())
        if not iso:
            continue

        availability = trainers[name]
        if not availability:
            continue

        years = frozenset(d.year for d in availability.keys())
        cache_key = (iso, years)
        if cache_key not in holiday_cache:
            try:
                holiday_cache[cache_key] = _holidays_lib.country_holidays(
                    iso, years=sorted(years)
                )
            except Exception:
                holiday_cache[cache_key] = {}
        country_holidays = holiday_cache[cache_key]

        for d in list(availability.keys()):
            if d in country_holidays:
                availability[d] = False


def parse_weightings(filepath: str) -> dict[str, int | float]:
    """Read the ``Weightings`` sheet into a ``{name: weight}`` dict.

    Returns ``{}`` if the workbook has no ``Weightings`` sheet. Weights are
    cast to ``int`` when possible, otherwise ``float``; uncoercible rows are
    skipped.
    """
    workbook = load_workbook(filepath, read_only=True, data_only=True)
    if "Weightings" not in workbook.sheetnames:
        return {}

    rows = list(workbook["Weightings"].iter_rows(values_only=True))
    if not rows:
        return {}

    name_idx, weight_idx = 0, 1
    for idx, cell in enumerate(rows[0]):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        if key == "name":
            name_idx = idx
        elif key == "weight":
            weight_idx = idx

    result: dict[str, int | float] = {}
    for row in rows[1:]:
        if not row:
            continue
        if name_idx >= len(row) or weight_idx >= len(row):
            continue
        raw_name, raw_weight = row[name_idx], row[weight_idx]
        if raw_name is None:
            continue
        name = str(raw_name).strip()
        if not name or raw_weight is None:
            continue
        try:
            result[name] = int(raw_weight)
        except (TypeError, ValueError):
            try:
                result[name] = float(raw_weight)
            except (TypeError, ValueError):
                continue
    return result


_LONG_HAUL_COUNTRIES_LOWER: frozenset[str] = frozenset(
    c.lower()
    for c in [
        "USA", "US", "United States",
        "India", "Singapore", "Australia",
        "China", "Japan", "Hong Kong",
        "South Africa", "Brazil", "New Zealand", "South Korea",
    ]
)
_TRAVEL_PENALTY = 10


def _is_long_haul(country: str | None) -> bool:
    if not country:
        return False
    return str(country).strip().lower() in _LONG_HAUL_COUNTRIES_LOWER


_LOCATION_FIELDS = {
    "name": "name",
    "country": "country",
    "demand": "demand",
    "french required": "french_required",
    "max parallel": "max_parallel",
    "max/week": "max_per_week",
    "max per week": "max_per_week",
}


def parse_locations(filepath: str) -> list[dict]:
    """Read the ``Locations`` sheet from a workbook.

    Returns:
        A list of location dicts, one per non-empty row. Each dict contains the
        keys ``name``, ``country``, ``demand``, ``french_required``,
        ``max_parallel``, and ``max_per_week``. Missing optional fields are
        filled with sensible defaults. Returns ``[]`` if the workbook has no
        ``Locations`` sheet.
    """
    workbook = load_workbook(filepath, read_only=True, data_only=True)
    if "Locations" not in workbook.sheetnames:
        return []
    worksheet = workbook["Locations"]

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    field_for_col: dict[int, str] = {}
    for col_idx, cell in enumerate(rows[0]):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        field = _LOCATION_FIELDS.get(key)
        if field is not None:
            field_for_col[col_idx] = field

    locations: list[dict] = []
    for row in rows[1:]:
        if not row:
            continue
        loc: dict = {}
        for col_idx, field in field_for_col.items():
            if col_idx >= len(row):
                continue
            value = row[col_idx]
            if field in ("name", "country"):
                loc[field] = str(value).strip() if value is not None else None
            elif field == "french_required":
                loc[field] = isinstance(value, str) and value.strip().lower() == "yes"
            elif field in ("demand", "max_parallel", "max_per_week"):
                if value is None:
                    loc[field] = 0 if field == "demand" else None
                else:
                    try:
                        loc[field] = int(value)
                    except (TypeError, ValueError):
                        loc[field] = 0 if field == "demand" else None

        if not loc.get("name"):
            continue

        loc.setdefault("country", None)
        loc.setdefault("demand", 0)
        loc.setdefault("french_required", False)
        loc.setdefault("max_parallel", None)
        loc.setdefault("max_per_week", None)
        locations.append(loc)
    return locations


def _parse_trainer_metadata(row: tuple, metadata_cols: dict[int, str]) -> dict:
    """Pull location / French / max-per-week from a trainer row."""
    info: dict = {
        "home_location": None,
        "french_speaking": False,
        "max_per_week": None,
    }
    for col_idx, field in metadata_cols.items():
        if col_idx >= len(row):
            continue
        value = row[col_idx]
        if field == "home_location":
            info["home_location"] = str(value).strip() if value is not None else None
        elif field == "french_speaking":
            info["french_speaking"] = isinstance(value, str) and value.strip().lower() == "yes"
        elif field == "max_per_week":
            if value is None:
                info["max_per_week"] = None
            else:
                try:
                    info["max_per_week"] = int(value)
                except (TypeError, ValueError):
                    info["max_per_week"] = None
    return info


def _is_available(cell) -> bool:
    """Return True iff a cell value is the literal string 'Yes' (case-sensitive, whitespace-tolerant)."""
    if not isinstance(cell, str):
        return False
    return cell.strip() == "Yes"


def generate_slots(
    dates: list[date],
    trainers: dict[str, dict[date, bool]],
    pattern: str = "mon-tue,thu-fri",
) -> list[tuple[date, date]]:
    """Return every valid two-day bootcamp slot.

    A slot ``(day1, day2)`` is returned iff:
      * ``day1`` and ``day2`` are both in ``dates``
      * ``day2 == day1 + 1 day``
      * ``(day1.weekday(), day2.weekday())`` matches an allowed pair from ``pattern``
      * At least two trainers have ``True`` for both days

    Args:
        dates: Sorted working dates from :func:`parse_availability`.
        trainers: Per-trainer availability map from :func:`parse_availability`.
        pattern: Comma-separated allowed day pairs (e.g. ``"mon-tue,thu-fri"``).

    Returns:
        A list of ``(day1, day2)`` tuples in ascending ``day1`` order.

    Raises:
        ValueError: If ``pattern`` is empty or contains an unknown day abbreviation.
    """
    allowed_pairs = _parse_pattern(pattern)
    date_set = set(dates)

    slots: list[tuple[date, date]] = []
    for day1 in dates:
        day2 = day1 + timedelta(days=1)
        if day2 not in date_set:
            continue
        if (day1.weekday(), day2.weekday()) not in allowed_pairs:
            continue
        if _count_available_both(trainers, day1, day2) >= 2:
            slots.append((day1, day2))
    return slots


def _parse_pattern(pattern: str) -> set[tuple[int, int]]:
    """Parse a day-pair pattern string into a set of ``(weekday1, weekday2)`` tuples."""
    if not pattern or not pattern.strip():
        raise ValueError("pattern is empty")
    pairs: set[tuple[int, int]] = set()
    for raw_token in pattern.split(","):
        token = raw_token.strip().lower()
        if not token:
            continue
        parts = [p.strip() for p in token.split("-")]
        if len(parts) != 2 or not all(parts):
            raise ValueError(f"Invalid pattern token: {raw_token!r}")
        try:
            w1, w2 = _DAY_ABBREV[parts[0]], _DAY_ABBREV[parts[1]]
        except KeyError as exc:
            raise ValueError(f"Unknown day abbreviation in token {raw_token!r}") from exc
        pairs.add((w1, w2))
    if not pairs:
        raise ValueError("pattern produced no day pairs")
    return pairs


def _count_available_both(
    trainers: dict[str, dict[date, bool]],
    day1: date,
    day2: date,
) -> int:
    """Count trainers available on both ``day1`` and ``day2``."""
    return sum(
        1
        for avail in trainers.values()
        if avail.get(day1, False) and avail.get(day2, False)
    )


def schedule_greedy(
    dates: list[date],
    trainers: dict[str, dict[date, bool]],
    slots: list[tuple[date, date]],
    config: dict | None = None,
) -> list[dict]:
    """Greedily assign two trainers per slot in trainer-priority order.

    A trainer is eligible for a slot iff they are available on both days and
    are not already booked on either day from an earlier slot in the same run.
    Slots that cannot be filled with two eligible trainers are skipped.

    When ``config`` provides a non-empty ``"experienced"`` set, an additional
    constraint applies: every assigned pair must include at least one
    experienced trainer. The pair selection bends the second slot only as much
    as needed to satisfy the constraint; the first slot is always the
    highest-priority eligible trainer.

    Args:
        dates: Working dates (accepted for signature compatibility, unused here).
        trainers: Per-trainer availability map. Iteration order defines priority.
        slots: Valid bootcamp slots in the order to consider them.
        config: Optional ``{"experienced": set[str], "trainees": set[str]}``.
            ``trainees`` is informational and not consulted. If ``config`` is
            ``None``, missing the key, or has an empty ``experienced`` set,
            no experience constraint is enforced.

    Returns:
        A list of ``{"slot": (d1, d2), "trainers": [name1, name2]}`` dicts,
        one per filled slot, in the input slot order.
    """
    experienced = _experienced_from_config(config)
    bookings: dict[date, set[str]] = {}
    schedule: list[dict] = []

    for slot in slots:
        day1, day2 = slot
        booked_d1 = bookings.get(day1, set())
        booked_d2 = bookings.get(day2, set())

        eligible = [
            name
            for name, avail in trainers.items()
            if avail.get(day1, False)
            and avail.get(day2, False)
            and name not in booked_d1
            and name not in booked_d2
        ]

        if len(eligible) < 2:
            continue

        pair = _pick_pair(eligible, experienced)
        if pair is None:
            continue

        schedule.append({"slot": slot, "trainers": pair})
        for name in pair:
            bookings.setdefault(day1, set()).add(name)
            bookings.setdefault(day2, set()).add(name)

    return schedule


def _experienced_from_config(config: dict | None) -> set[str]:
    """Extract the experienced-trainer set from ``config``. Empty set means no constraint."""
    if not config:
        return set()
    experienced = config.get("experienced") or set()
    return set(experienced)


def _pick_pair(eligible: list[str], experienced: set[str]) -> list[str] | None:
    """Choose the two trainers for a slot, honoring experience rules if any.

    Returns ``None`` if the experience constraint can't be satisfied.
    """
    if not experienced:
        return eligible[:2]

    first = eligible[0]
    if first in experienced:
        return [first, eligible[1]]

    # First is not experienced — the partner must be.
    for candidate in eligible[1:]:
        if candidate in experienced:
            return [first, candidate]
    return None


def schedule_optimal(
    dates: list[date],
    trainers: dict[str, dict[date, bool]],
    slots: list[tuple[date, date]],
    config: dict | None = None,
    trainer_info: dict | None = None,
    weekly_caps: dict | None = None,
    locations: list[dict] | None = None,
    weightings: dict[str, int | float] | None = None,
) -> list[dict]:
    """Assign two trainers per slot using CP-SAT to maximise filled bootcamps.

    Args:
        dates: Working dates (used only for indirect derivation of per-day
            booking constraints).
        trainers: Per-trainer availability map. Iteration order defines the
            output order of names within each pair.
        slots: Candidate bootcamp slots.
        config: Optional ``{"experienced": set[str], "trainees": set[str]}``.
            When ``experienced`` is non-empty, every filled slot must include
            at least one name from it.
        trainer_info: Optional per-trainer metadata from :func:`parse_availability`.
            When provided, any trainer with ``max_per_week`` set has a per-week
            bootcamp cap of ``max_per_week // 2`` applied.
        weekly_caps: Optional ``{(iso_year, iso_week): int}`` map. When provided,
            the total bootcamps in any listed week cannot exceed its cap.
        locations: Optional list of location dicts from :func:`parse_locations`.
            When provided and non-empty, each filled slot is assigned to exactly
            one location, capped at the location's ``demand``, and French-required
            locations require at least one French-speaking trainer in the pair.
            Each result entry then carries a ``"location"`` key.
        weightings: Optional ``{name: weight}`` mapping. When non-empty, the
            objective gains a soft bonus proportional to assigned-trainer weight,
            so ties between equally-large schedules favour higher-weighted
            trainers. A soft travel penalty is also applied to any slot assigned
            to a long-haul location (small effect; deterministic across solves).

    Returns:
        A list of ``{"slot": (d1, d2), "trainers": [name1, name2]}`` dicts,
        optionally with a ``"location"`` key when ``locations`` is provided.
    """
    experienced = _experienced_from_config(config)
    trainer_names = list(trainers.keys())
    locations_list = list(locations) if locations else []

    model = cp_model.CpModel()

    # x[(name, slot_idx)] — trainer assigned to slot? Only created when available on both days.
    x: dict[tuple[str, int], cp_model.IntVar] = {}
    for slot_idx, (day1, day2) in enumerate(slots):
        for name in trainer_names:
            avail = trainers[name]
            if avail.get(day1, False) and avail.get(day2, False):
                x[(name, slot_idx)] = model.new_bool_var(f"x_{name}_{slot_idx}")

    # y[slot_idx] — slot filled?
    y = {
        slot_idx: model.new_bool_var(f"y_{slot_idx}")
        for slot_idx in range(len(slots))
    }

    # Constraint: pair size — sum of trainers on a slot is either 0 or 2.
    for slot_idx in range(len(slots)):
        assigned = [x[(name, slot_idx)] for name in trainer_names if (name, slot_idx) in x]
        if assigned:
            model.add(sum(assigned) == 2 * y[slot_idx])
        else:
            model.add(y[slot_idx] == 0)

    # Constraint: no double-booking per (trainer, day).
    day_to_slot_indices: dict[date, list[int]] = {}
    for slot_idx, (day1, day2) in enumerate(slots):
        day_to_slot_indices.setdefault(day1, []).append(slot_idx)
        day_to_slot_indices.setdefault(day2, []).append(slot_idx)
    for name in trainer_names:
        for day, slot_indices in day_to_slot_indices.items():
            vars_for_day = [
                x[(name, slot_idx)] for slot_idx in slot_indices if (name, slot_idx) in x
            ]
            if len(vars_for_day) > 1:
                model.add(sum(vars_for_day) <= 1)

    # Constraint: experience rule (only when enforcing).
    if experienced:
        for slot_idx in range(len(slots)):
            exp_vars = [
                x[(name, slot_idx)]
                for name in trainer_names
                if name in experienced and (name, slot_idx) in x
            ]
            if exp_vars:
                model.add(sum(exp_vars) >= y[slot_idx])
            else:
                # No experienced trainer available — slot cannot be filled.
                model.add(y[slot_idx] == 0)

    # Group slots by ISO week of their first day — used by both cap families.
    slots_by_week: dict[tuple[int, int], list[int]] = {}
    for slot_idx, (day1, _) in enumerate(slots):
        iso_year, iso_week, _ = day1.isocalendar()
        slots_by_week.setdefault((iso_year, iso_week), []).append(slot_idx)

    # Constraint: per-trainer weekly bootcamp cap.
    if trainer_info:
        for name in trainer_names:
            info = trainer_info.get(name) or {}
            max_per_week = info.get("max_per_week")
            if max_per_week is None:
                continue
            max_bootcamps = max_per_week // 2
            for week_key, week_slot_indices in slots_by_week.items():
                vars_for_week = [
                    x[(name, slot_idx)]
                    for slot_idx in week_slot_indices
                    if (name, slot_idx) in x
                ]
                if vars_for_week:
                    model.add(sum(vars_for_week) <= max_bootcamps)

    # Constraint: per-week total bootcamp cap.
    if weekly_caps:
        for week_key, cap in weekly_caps.items():
            week_slot_indices = slots_by_week.get(week_key, [])
            if not week_slot_indices:
                continue
            model.add(sum(y[slot_idx] for slot_idx in week_slot_indices) <= cap)

    # Location assignment: one location per filled slot, capped by demand.
    loc_var: dict[tuple[int, int], cp_model.IntVar] = {}
    french_speakers = {
        name for name in trainer_names
        if trainer_info and trainer_info.get(name, {}).get("french_speaking", False)
    }
    if locations_list:
        for slot_idx in range(len(slots)):
            for loc_idx in range(len(locations_list)):
                loc_var[(slot_idx, loc_idx)] = model.new_bool_var(f"loc_{slot_idx}_{loc_idx}")

        # Each filled slot is assigned to exactly one location.
        for slot_idx in range(len(slots)):
            slot_loc_vars = [loc_var[(slot_idx, l)] for l in range(len(locations_list))]
            model.add(sum(slot_loc_vars) == y[slot_idx])

        # Per-location demand cap.
        for loc_idx, location in enumerate(locations_list):
            demand = int(location.get("demand", 0) or 0)
            model.add(
                sum(loc_var[(s, loc_idx)] for s in range(len(slots))) <= demand
            )

        # French-speaking requirement on French-required locations.
        for loc_idx, location in enumerate(locations_list):
            if not location.get("french_required"):
                continue
            for slot_idx in range(len(slots)):
                french_x = [
                    x[(name, slot_idx)]
                    for name in trainer_names
                    if name in french_speakers and (name, slot_idx) in x
                ]
                if french_x:
                    model.add(loc_var[(slot_idx, loc_idx)] <= sum(french_x))
                else:
                    model.add(loc_var[(slot_idx, loc_idx)] == 0)

    # Objective: maximise filled slots (hard), then weighting bonus, then minus travel penalty.
    weightings_map = dict(weightings) if weightings else {}
    weight_term = sum(
        weightings_map.get(name, 0) * x[(name, slot_idx)]
        for (name, slot_idx) in x
    )
    travel_long_haul = [
        loc_var[(slot_idx, loc_idx)]
        for slot_idx in range(len(slots))
        for loc_idx, location in enumerate(locations_list)
        if _is_long_haul(location.get("country"))
    ]
    # Each filled slot has 2 trainers, so penalise 2 * TRAVEL_PENALTY per long-haul (slot, loc).
    travel_term = sum(2 * _TRAVEL_PENALTY * v for v in travel_long_haul)

    # BIG must exceed any possible swing in soft terms so filling one more slot
    # is strictly better than any reshuffle.
    max_weight_swing = 2 * sum(int(w) for w in weightings_map.values()) if weightings_map else 0
    max_travel_swing = 2 * _TRAVEL_PENALTY * len(travel_long_haul)
    big = 1 + max_weight_swing + max_travel_swing

    model.maximize(big * sum(y.values()) + weight_term - travel_term)

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 30
    status = solver.solve(model)

    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        return []

    schedule: list[dict] = []
    for slot_idx, slot in enumerate(slots):
        if solver.value(y[slot_idx]) != 1:
            continue
        assigned_names = [
            name
            for name in trainer_names
            if (name, slot_idx) in x and solver.value(x[(name, slot_idx)]) == 1
        ]
        entry: dict = {"slot": slot, "trainers": assigned_names}
        if locations_list:
            for loc_idx, location in enumerate(locations_list):
                if solver.value(loc_var[(slot_idx, loc_idx)]) == 1:
                    entry["location"] = location["name"]
                    break
        schedule.append(entry)
    return schedule
